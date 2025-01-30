import telebot
import openpyxl
import time
import yfinance as yf
from YFinance_API import fetch_stock_data
from YFinance_API import Stocks
from datetime import datetime

api_token = "BOT_API_TOKEN"

bot = telebot.TeleBot(api_token, parse_mode=None) #bot token

# /start and /help handler
@bot.message_handler(commands=['help','start'])
def welcome(message):
    bot.send_message(message.chat.id, """Hi User, call the "/stock" command to get informatons on a stock.""")

# Step 1: /stocks command to initiate the stock query
@bot.message_handler(commands=['stocks', 'stock'])
def ask_for_stock_name(message):
    bot.send_message(message.chat.id, "Please enter the stock symbol (e.g., AAPL for Apple):")
    bot.register_next_step_handler(message, get_stock_info)

# Step 2: Function to process the user’s input and retrieve stock information
def get_stock_info(message):
    # Insert a try and except structure to handle exceptions without stopping the bot 
    stock_symbol = message.text.upper()
    
    # Check whether the user is blocked
    if is_blocked(message.from_user.id):
        bot.send_message(message.chat.id, "Wait at least 30 seconds between a request and another. Please try again later.")
    else:
        # Update the Excel file
        update_excel_file(message.from_user.id, message.from_user.username, time.time())
        try:
            stock_info = fetch_stock_data(stock_symbol)
        except:
            bot.send_message(message.chat.id, f"Sorry, I couldn't find information for the stock symbol '{stock_symbol}'.\nMake sure the format and the stock's name are correct and try again.")
        else:
            bot.send_message(message.chat.id, f"Here is the information for {stock_symbol}:\n\n{stock_info}")

# Function to update the Excel file and limit the number of queries per minute
def update_excel_file(user_id, user_name = "DefaultUsername", current_query_time = 0):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook("/Users/lorenzogalluzzi/Documents/programming/Telebot/Telebot_Database.xlsx")
    sheet = workbook.active

    # Find the next empty row
    next_row = sheet.max_row + 1
    
    # Format the timestamp
    formatted_time = datetime.fromtimestamp(current_query_time).strftime('%d/%m/%Y %H:%M:%S')

    # check if the UserID has already been registered
    seen = False
    current_queries = 0
    for i in range(1, next_row):
        
        if sheet[f'A{i}'].value == user_id:

            current_queries = int(sheet[f'C{i}'].value)
            sheet[f'C{i}'] = current_queries + 1
            sheet[f'D{i}'] = formatted_time
            seen = True    
            break
    
    # Add user ID, user name, timestamp, and the number of queries
    if not seen:
        sheet[f'A{next_row}'] = user_id
        sheet[f'B{next_row}'] = user_name
        sheet[f'C{next_row}'] = 1
        sheet[f'D{next_row}'] = formatted_time

    # Save the workbooks
    workbook.save("/Users/lorenzogalluzzi/Documents/programming/Telebot/Telebot_Database.xlsx")

# Determine whether an user is blocked or not
def is_blocked(user_id):
    workbook = openpyxl.load_workbook("/Users/lorenzogalluzzi/Documents/programming/Telebot/Telebot_Database.xlsx")
    try:    
        sheet = workbook.active
        next_row = sheet.max_row + 1
        last_query_time = 0
        for i in range(1, next_row):
            if sheet[f'A{i}'].value == user_id:
                
                # Convert the last query time to a timestamp
                last_query_time = datetime.strptime(sheet[f'D{i}'].value, '%d/%m/%Y %H:%M:%S').timestamp()
                
                # Check if the last query was made more than a minute ago
                if time.time() - last_query_time > 30:
                    
                    return False
                return True 
        return False
    finally:
        workbook.close()
# Start the bot
bot.polling()
